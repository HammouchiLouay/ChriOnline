from __future__ import annotations

from pathlib import Path
import re

import openpyxl


def _is_number(v: object) -> bool:
    return isinstance(v, (int, float))


def _round2(x: float) -> float:
    return float(f"{x:.2f}")


def update_workbook(path: Path) -> int:
    """
    Increases estimation and effective time values across the workbook.

    Rules:
    - Estimated time is scaled up (more time than currently).
    - Effective time is scaled up and ends slightly above estimated.
    - Keeps the sheet layout; only updates numeric cells under the relevant headers.
    """

    wb = openpyxl.load_workbook(path)

    header_re_est = re.compile(r"estim", re.IGNORECASE)
    header_re_eff = re.compile(r"(temps\s*effectif|effectif|temps\s*r[eé]el)", re.IGNORECASE)

    changed_rows = 0

    for ws in wb.worksheets:
        est_col = None
        eff_col = None
        header_row = None

        # Find header positions in the top-left area.
        for r in range(1, min(30, ws.max_row) + 1):
            for c in range(1, min(30, ws.max_column) + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                s = v.strip()
                if est_col is None and header_re_est.search(s):
                    est_col = c
                    header_row = r
                if eff_col is None and header_re_eff.search(s):
                    eff_col = c
                    header_row = r
            if est_col is not None and eff_col is not None and header_row is not None:
                break

        if est_col is None or eff_col is None or header_row is None:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            first = ws.cell(r, 1).value
            if isinstance(first, str):
                fu = first.strip().upper()
                if fu in ("TOTAL", "TOTALE", "TOTALE (FONCTIONNEL)"):
                    continue
            est_cell = ws.cell(r, est_col)
            eff_cell = ws.cell(r, eff_col)
            est_v = est_cell.value
            eff_v = eff_cell.value

            if not _is_number(est_v) and not _is_number(eff_v):
                continue

            old_est = float(est_v) if _is_number(est_v) else None
            old_eff = float(eff_v) if _is_number(eff_v) else None

            # If one is missing, infer conservatively to keep ratio behaviour stable.
            if old_est is None and old_eff is not None:
                old_est = old_eff * 0.92
            if old_eff is None and old_est is not None:
                old_eff = old_est * 1.03

            # Modest increase ("un peu") — avoid re-running on already bumped rows.
            new_est = _round2(max(0.05, old_est * 1.08))
            # Effective typically slightly above estimated.
            new_eff = _round2(max(old_eff * 1.10, new_est * 1.03))

            if _is_number(est_v):
                est_cell.value = new_est
            if _is_number(eff_v):
                eff_cell.value = new_eff

            changed_rows += 1

    wb.save(path)
    return changed_rows


def main() -> None:
    path = Path(r"c:\Users\Legion\Downloads\ChriOnline\Gestion des utilisateurs.xlsx")
    changed = update_workbook(path)
    print(f"Updated rows: {changed}")


if __name__ == "__main__":
    main()

