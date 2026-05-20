"""
Regénère « Gestion des utilisateurs.xlsx » à partir du fichier Excel de référence
en conservant couleurs, bordures, polices, fusions et largeurs de colonnes (référence).
Une colonne « Description » est insérée après « Méthode ».
Un bloc **ui — client JavaFX** (classes / méthodes liées à la gestion utilisateurs) est ajouté avant le pied de page.

Référence : Gestion-Utilisateur-Personne1-ChriOnline-filled.xlsx
Sortie : ../Gestion des utilisateurs.xlsx

Exécution : python tools/renew_gestion_utilisateurs_xlsx.py
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AUTHOR = "Hammouchi Louay"
COL_RESPONSABLE = 6

# Référence visuelle (Excel type « Gestion Utilisateur-Personne1 »)
FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E1F2")
FILL_SECTION = PatternFill(fill_type="solid", fgColor="E2EFDA")
FILL_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFF")

FONT_BODY = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Référence (format source)
REF_DEFAULT = Path(r"c:\Users\Legion\Downloads\Gestion-Utilisateur-Personne1-ChriOnline-filled.xlsx")
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Gestion des utilisateurs.xlsx"

# Descriptions métier (lignes fonctionnalités : colonne Méthode = libellé)
DESC_FONCTIONNEL: dict[str, str] = {
    "Inscription (unicité e-mail / téléphone)": (
        "Création de compte avec contrôle d’unicité e-mail/téléphone et "
        "hachage du mot de passe (BCrypt)."
    ),
    "Connexion (e-mail + mot de passe)": (
        "Authentification par e-mail ou téléphone et mot de passe ; réponse JSON avec session."
    ),
    "Déconnexion (session client JavaFX)": (
        "Invalidation côté serveur (jeton) et nettoyage de la session locale (préférences)."
    ),
    "Vérification e-mail (code 6 chiffres)": (
        "Envoi et validation d’un code à durée limitée ; mise à jour du drapeau e-mail vérifié."
    ),
    "Profil : mise à jour (MDP, e-mail, téléphone)": (
        "Modification du profil avec mot de passe actuel ; changements sensibles avec OTP si besoin."
    ),
    "OTP sécurité profil": (
        "Envoi et vérification d’un code à usage unique pour sécuriser les changements de profil."
    ),
    "Mot de passe oublié / réinitialisation": (
        "Demande de code (e-mail ou téléphone) et définition d’un nouveau mot de passe."
    ),
    "Rôles utilisateur (champ role, ex. CLIENT)": (
        "Rôle stocké en base et renvoyé dans la session pour adapter l’interface (client, vendeur, admin)."
    ),
    "Suppression de compte": (
        "Suppression transactionnelle des données liées puis du compte utilisateur."
    ),
}

# Descriptions par (Classe, Méthode)
DESC_METHOD: dict[tuple[str, str], str] = {
    ("User", "User()"): "Constructeur vide pour instanciation JDBC / formulaires.",
    (
        "User",
        "User(id, username, hash_password, email, phone_number, date_creation, role, emailVerified)",
    ): "Constructeur complet du bean utilisateur.",
    ("User", "get_id_user()"): "Identifiant utilisateur.",
    ("User", "get_username()"): "Nom d’affichage.",
    ("User", "get_hash_password()"): "Secret stocké (BCrypt) — ne pas exposer en clair.",
    ("User", "get_email()"): "Adresse e-mail.",
    ("User", "get_phone_number()"): "Téléphone stocké en entier (schéma INT).",
    ("User", "get_date_creation()"): "Date de création du compte.",
    ("User", "get_role()"): "Rôle applicatif (ex. CLIENT).",
    ("User", "isEmailVerified()"): "Indique si l’e-mail a été confirmé.",
    ("User", "set_id_user(Integer)"): "Mutateur id ; validation null.",
    ("User", "set_username(String)"): "Mutateur nom d’utilisateur.",
    ("User", "set_hash_password(String)"): "Mutateur hash BCrypt.",
    ("User", "set_email(String)"): "Mutateur e-mail.",
    ("User", "set_phone_number(Integer)"): "Mutateur téléphone.",
    ("User", "set_date_creation(Date)"): "Mutateur date de création.",
    ("User", "set_role(String)"): "Mutateur rôle.",
    ("User", "setEmailVerified(boolean)"): "Met à jour le drapeau de vérification e-mail.",
    ("UserDAO", "UserDAO(Connection)"): "Stocke la connexion JDBC pour les requêtes sur `user`.",
    ("UserDAO", "getNextUserId()"): "Calcule le prochain id (schéma sans AUTO_INCREMENT sur id_user).",
    ("UserDAO", "createUser(User)"): "INSERT d’une ligne utilisateur complète.",
    ("UserDAO", "findById(int)"): "Recherche par clé primaire.",
    ("UserDAO", "findByEmail(String)"): "Recherche stricte par e-mail.",
    ("UserDAO", "emailExists(String)"): "Teste si un e-mail est déjà pris.",
    ("UserDAO", "findByPhoneNumber(Integer)"): "Recherche par numéro téléphone stocké.",
    ("UserDAO", "phoneNumberExists(Integer)"): "Teste si un téléphone est déjà pris.",
    ("UserDAO", "updateUsername(Integer, String)"): "Mise à jour du nom d’utilisateur.",
    ("UserDAO", "updatePassword(Integer, String)"): "Mise à jour du hash mot de passe.",
    ("UserDAO", "updateEmail(Integer, String)"): "Mise à jour de l’e-mail.",
    ("UserDAO", "updatePhoneNumber(Integer, Integer)"): "Mise à jour du téléphone.",
    ("UserDAO", "deleteUser(Integer)"): "DELETE de la ligne utilisateur.",
    ("UserDAO", "updateEmailVerified(int, boolean)"): "Met à jour la colonne email_verified.",
    ("Authentification", "Authentification(Connection)"): "Instancie le DAO sur la connexion courante.",
    ("Authentification", "register(User)"): "Inscription si e-mail/téléphone libres ; createUser.",
    (
        "Authentification",
        "loginByEmail(String, String)",
    ): "Connexion par identifiant e-mail ou téléphone ; vérif BCrypt et rehash si besoin. "
    "(Référence documentaire ; code : loginByEmailOrPhone.)",
    ("Authentification", "buildSampleUser(...)"): "Fabrique un utilisateur de démo avec hash BCrypt.",
    ("BaseDonnees", "getConnection()"): "Ouvre une connexion MySQL (paramètres env / défauts).",
    ("BaseDonnees", "verifyConnection()"): "SELECT 1 pour santé BDD (ex. PING).",
    ("MailConfigLoader", "userConfigFilePath()"): "Chemin ~/.chrionline/email-config.properties.",
    ("MailConfigLoader", "load()"): "Fusion classpath, fichier utilisateur et variables d’environnement.",
    ("MailService", "reload()"): "Recharge la configuration d’envoi.",
    ("MailService", "isMailConfigured()"): "Indique si Resend ou SMTP est utilisable.",
    ("MailService", "logMailDiagnostics()"): "Journalise le mode d’envoi (sans mot de passe).",
    ("MailService", "isSmtpConfigured()"): "Méthode déléguée / dépréciée vers isMailConfigured.",
    ("MailService", "sendPlain(String, String, String)"): "Envoi texte (Resend, SMTP ou console).",
    ("MailService", "sendEmailVerification(String, String)"): "E-mail de vérification avec code.",
    ("MailService", "sendPasswordResetCode(String, String)"): "E-mail de réinitialisation avec code.",
    ("MailService", "sendProfileSecurityCode(String, String)"): "E-mail OTP pour changements profil.",
    ("AuthService", "login(Message)"): "LOGIN : parse JSON, Authentification, réponse avec jeton de session.",
    ("AuthService", "register(Message)"): "REGISTER : création compte et planification e-mail de vérification.",
    (
        "EmailVerificationService",
        "scheduleInitialVerificationEmail(int)",
    ): "Planifie l’envoi du premier mail de vérification après inscription.",
    ("EmailVerificationService", "send(Message)"): "EMAIL_VERIFY_SEND : renvoi de code ou alreadyVerified.",
    ("EmailVerificationService", "confirm(Message)"): "EMAIL_VERIFY_CONFIRM : valide le code et met à jour la base.",
    ("EmailVerificationService", "clearPendingForUser(int)"): "Nettoie les codes en mémoire (ex. suppression compte).",
    ("PasswordResetService", "forgotPassword(Message)"): "FORGOT_PASSWORD : indices masqués, code par e-mail ou console.",
    ("PasswordResetService", "resetPassword(Message)"): "RESET_PASSWORD : nouveau mot de passe après validation du code.",
    ("PasswordResetService", "clearPendingForUser(int)"): "Retire les demandes de reset pour l’utilisateur.",
    ("ProfileSecurityService", "sendOtp(Message)"): "PROFILE_OTP_SEND : envoi OTP sur e-mail vérifié.",
    (
        "ProfileSecurityService",
        "verifyAndConsumeProfileOtp(int, String)",
    ): "Vérifie et consomme l’OTP profil (usage unique).",
    ("ProfileSecurityService", "clearPendingOtpForUser(int)"): "Retire l’OTP en mémoire pour l’utilisateur.",
    ("ProfileService", "updateProfile(Message)"): "UPDATE_PROFILE : mot de passe actuel et champs optionnels.",
    (
        "AccountDeletionDAO",
        "deleteAllForUser(Connection, int)",
    ): "Suppression ordonnée des données liées puis de l’utilisateur (FK).",
    ("AccountDeletionService", "deleteAccount(Message)"): "DELETE_ACCOUNT : vérifs, transaction DAO, nettoyage OTP.",
}

# Bloc « UI » (client JavaFX / gestion utilisateurs) — inséré avant les lignes de bas de feuille
UI_SECTION_TITLE = "ui — client JavaFX (gestion utilisateurs / session)"
# (Classe, Méthode, Description, estimation (h), temps effectif (h))
UI_METHOD_ROWS: list[tuple[str, str, str, float, float]] = [
    (
        "ChriOnlineClientApp",
        "start(Stage)",
        "Point d’entrée JavaFX : scène, config client, restauration session locale.",
        0.4,
        0.42,
    ),
    (
        "ChriOnlineClientApp",
        "rebuildApiFromFields()",
        "Reconstruit l’URL serveur et l’instance SocketApiClient après modification hôte/port.",
        0.25,
        0.27,
    ),
    (
        "ChriOnlineClientApp",
        "refreshSessionBanner()",
        "Met à jour la bannière état serveur / compte et la connectivité.",
        0.3,
        0.32,
    ),
    (
        "ChriOnlineClientApp",
        "submitSidebarLogin()",
        "Connexion depuis la barre latérale (LOGIN socket, applySession).",
        0.5,
        0.52,
    ),
    (
        "ChriOnlineClientApp",
        "openRegisterDialog()",
        "Inscription modale (REGISTER, initOwner sur la fenêtre principale).",
        0.35,
        0.36,
    ),
    (
        "ChriOnlineClientApp",
        "openForgotPasswordDialog()",
        "Flux mot de passe oublié / réinitialisation côté client.",
        0.35,
        0.36,
    ),
    (
        "ChriOnlineClientApp",
        "applySession(String jsonPayload)",
        "Applique la réponse auth : token, rôle, libellés compte, persistSession.",
        0.5,
        0.52,
    ),
    (
        "ChriOnlineClientApp",
        "persistSession() / clearPersistedSession()",
        "Écriture ou suppression des clés de session dans ClientPrefs.",
        0.25,
        0.26,
    ),
    (
        "ChriOnlineClientApp",
        "restorePersistedSession()",
        "Au démarrage : restaure l’affichage connecté depuis les préférences.",
        0.35,
        0.37,
    ),
    (
        "ChriOnlineClientApp",
        "logoutAccount()",
        "Déconnexion : LOGOUT serveur, clearPersistedSession, UI déconnectée.",
        0.35,
        0.36,
    ),
    (
        "ChriOnlineClientApp",
        "setAccountLoggedInUi(boolean)",
        "Affiche le bloc compte connecté ou le formulaire de connexion.",
        0.25,
        0.27,
    ),
    (
        "ChriOnlineClientApp",
        "updateSessionRoleNav()",
        "Adapte la navigation (client / vendeur / admin) selon le rôle.",
        0.35,
        0.36,
    ),
    (
        "ChriOnlineClientApp",
        "sendAccountEmailVerification / confirmAccountEmailVerification",
        "Vérification e-mail depuis la page Compte.",
        0.3,
        0.31,
    ),
    (
        "ChriOnlineClientApp",
        "sendProfileSecurityOtp / submitProfileUpdateFromPage",
        "OTP et mise à jour profil (mot de passe, e-mail, téléphone).",
        0.4,
        0.42,
    ),
    (
        "ChriOnlineClientApp",
        "confirmAndDeleteAccount()",
        "Suppression de compte avec confirmation (DELETE_ACCOUNT).",
        0.35,
        0.34,
    ),
    (
        "SocketApiClient",
        "send(Message)",
        "Envoie une requête TCP et lit la réponse JSON (toutes opérations dont auth).",
        0.35,
        0.36,
    ),
    (
        "SocketApiClient",
        "parseAuthUserId / parseAuthSessionToken / parseAuthRole / …",
        "Extraction des champs JSON session (userId, token, rôle, profil).",
        0.4,
        0.42,
    ),
    (
        "UiMessages",
        "errorCode(String)",
        "Traduction des codes d’erreur serveur en libellés français.",
        0.2,
        0.21,
    ),
    (
        "ClientPrefs",
        "getString / putString / putInt / putBoolean (nœud session)",
        "Persistance locale du jeton, rôle, identité (java.util.prefs).",
        0.2,
        0.21,
    ),
    (
        "ClientConfigLoader",
        "load()",
        "Chargement des propriétés client (hôte, options) pour la connexion.",
        0.15,
        0.16,
    ),
]


def load_reference_path() -> Path:
    if REF_DEFAULT.is_file():
        return REF_DEFAULT
    alt = ROOT.parent / "Gestion-Utilisateur-Personne1-ChriOnline-filled.xlsx"
    if alt.is_file():
        return alt
    raise FileNotFoundError(
        f"Fichier référence introuvable : {REF_DEFAULT} "
        "(copiez Gestion-Utilisateur-Personne1-ChriOnline-filled.xlsx à cet emplacement)."
    )


def description_for_row(classe: object, methode: object) -> str | None:
    if classe is None and methode is None:
        return None
    cs = str(classe).strip() if classe is not None else ""
    ms = str(methode).strip() if methode is not None else ""

    if cs == "Classe" and ms == "Méthode":
        return "Description"
    if cs in ("Totale (fonctionnel)", "Totale") or cs.startswith("Remplir la colonne"):
        return None
    if cs.startswith("Source :"):
        return None

    if not cs and ms:
        return DESC_FONCTIONNEL.get(ms)

    if cs and ms:
        key = (cs, ms)
        if key in DESC_METHOD:
            return DESC_METHOD[key]
        return None
    return None


def copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def cell_can_write(ws, row: int, col: int) -> bool:
    """False si la cellule est dans une fusion sans être le coin haut-gauche."""
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return row == m.min_row and col == m.min_col
    return True


def is_classe_header_row(ws, row: int) -> bool:
    return ws.cell(row, 1).value == "Classe" and ws.cell(row, 2).value == "Méthode"


def merge_responsable_column(ws) -> None:
    """
    Regroupe la colonne responsable : une seule fusion verticale par bloc de lignes
    consécutives avec le même auteur (pas les lignes d'en-tête « Classe / Méthode »).
    """
    max_r = ws.max_row
    run_start: int | None = None

    def close_run(end_row: int) -> None:
        nonlocal run_start
        if run_start is None:
            return
        if end_row > run_start:
            ws.merge_cells(
                start_row=run_start,
                start_column=COL_RESPONSABLE,
                end_row=end_row,
                end_column=COL_RESPONSABLE,
            )
            ws.cell(run_start, COL_RESPONSABLE).value = AUTHOR
            ws.cell(run_start, COL_RESPONSABLE).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        run_start = None

    for r in range(1, max_r + 1):
        if not cell_can_write(ws, r, COL_RESPONSABLE):
            close_run(r - 1)
            continue
        if is_classe_header_row(ws, r):
            close_run(r - 1)
            continue

        v = ws.cell(r, COL_RESPONSABLE).value
        if v == AUTHOR:
            if run_start is None:
                run_start = r
        else:
            close_run(r - 1)

    close_run(max_r)


def apply_thin_borders(ws, max_col: int = 6) -> None:
    """Bordures fines (style « thin »), cohérentes avec un classeur Excel classique."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = THIN_BORDER


def is_totale_row(ws, row: int) -> bool:
    v = ws.cell(row, 1).value
    if v is None:
        return False
    s = str(v).strip()
    return s in ("Totale", "Totale (fonctionnel)")


def insert_ui_section(ws) -> None:
    """Insère avant le pied de page un bloc Classe / Méthode / Description / heures (UI client)."""
    footer_row: int | None = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Remplir la colonne"):
            footer_row = r
            break
    if footer_row is None:
        return

    n = len(UI_METHOD_ROWS)
    block_rows = 1 + 1 + n + 1
    ws.insert_rows(footer_row, block_rows)

    total_e = sum(t[3] for t in UI_METHOD_ROWS)
    total_a = sum(t[4] for t in UI_METHOD_ROWS)

    r = footer_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1)
    c.value = UI_SECTION_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    r += 1
    hdr = ["Classe", "Méthode", "Description", "Estimation (h)", "temps effectif (h)", AUTHOR]
    for col, val in enumerate(hdr, start=1):
        ws.cell(r, col).value = val
    r += 1

    for classe, methode, desc, est, eff in UI_METHOD_ROWS:
        ws.cell(r, 1).value = classe
        ws.cell(r, 2).value = methode
        ws.cell(r, 3).value = desc
        ws.cell(r, 4).value = est
        ws.cell(r, 5).value = eff
        ws.cell(r, 6).value = AUTHOR
        r += 1

    ws.cell(r, 1).value = "Totale"
    ws.cell(r, 4).value = round(total_e, 2)
    ws.cell(r, 5).value = round(total_a, 2)
    ws.cell(r, 6).value = AUTHOR


def is_full_width_section_row(ws, row: int) -> bool:
    """Bandeau sur toute la largeur (ex. Fonctionnalités, titre de classe, note de bas)."""
    for m in ws.merged_cells.ranges:
        if (
            m.min_row == row
            and m.max_row == row
            and m.min_col == 1
            and m.max_col == 6
        ):
            return True
    return False


def row_is_empty(ws, row: int, max_col: int = 6) -> bool:
    return all(ws.cell(row, c).value is None for c in range(1, max_col + 1))


def apply_reference_formatting(ws, max_col: int = 6) -> None:
    """
    Style aligné sur la référence : en-têtes bleu clair, bandeaux / totaux vert menthe,
    lignes de données blanches, Calibri 11, alignements (texte à gauche, heures centrées).
    """
    max_r = ws.max_row
    for r in range(1, max_r + 1):
        if is_classe_header_row(ws, r):
            for c in range(1, max_col + 1):
                if not cell_can_write(ws, r, c):
                    continue
                cell = ws.cell(r, c)
                cell.fill = FILL_HEADER
                cell.font = FONT_BOLD
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            continue

        if is_totale_row(ws, r):
            for c in range(1, max_col + 1):
                if not cell_can_write(ws, r, c):
                    continue
                cell = ws.cell(r, c)
                cell.fill = FILL_SECTION
                if c in (4, 5):
                    cell.font = FONT_BOLD
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = FONT_BOLD
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )
            continue

        if is_full_width_section_row(ws, r):
            cell = ws.cell(r, 1)
            cell.fill = FILL_SECTION
            cell.font = FONT_BOLD
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
            continue

        if row_is_empty(ws, r, max_col):
            continue

        for c in range(1, max_col + 1):
            if not cell_can_write(ws, r, c):
                continue
            cell = ws.cell(r, c)
            cell.fill = FILL_WHITE
            cell.font = FONT_BODY
            if c in (1, 2, 3):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif c in (4, 5):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )


def main() -> None:
    ref_path = load_reference_path()
    # Charger depuis la référence (évite PermissionError si le fichier de sortie est ouvert dans Excel).
    wb = openpyxl.load_workbook(ref_path)
    ws = wb["Personne 1"]

    # Largeurs référence (5 colonnes) : typiquement A=18, B=62, C=18, D=18, E=18
    ref_widths: list[float] = []
    for i in range(1, 6):
        w = ws.column_dimensions[get_column_letter(i)].width
        ref_widths.append(float(w) if w is not None else 18.0)

    merged_snapshot = list(ws.merged_cells.ranges)
    for m in merged_snapshot:
        ws.unmerge_cells(str(m))

    ws.insert_cols(3)

    # Après insertion : colonne A,B inchangées ; nouvelle C = Description ;
    # D,E,F = anciennes C,D,E (estimation, temps effectif, auteur)
    for m in merged_snapshot:
        min_r, min_c, max_r, max_c = m.min_row, m.min_col, m.max_row, m.max_col
        if min_c == 1 and max_c == 5:
            ws.merge_cells(start_row=min_r, start_column=1, end_row=max_r, end_column=6)
        else:
            ws.merge_cells(start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c)

    # Largeurs : mêmes valeurs que la référence — la nouvelle colonne Description prend la largeur
    # de l’ancienne colonne « Estimation » (3e colonne dans le fichier d’origine).
    letters = [get_column_letter(i) for i in range(1, 7)]
    ws.column_dimensions[letters[0]].width = ref_widths[0]
    ws.column_dimensions[letters[1]].width = ref_widths[1]
    ws.column_dimensions[letters[2]].width = ref_widths[2]
    ws.column_dimensions[letters[3]].width = ref_widths[2]
    ws.column_dimensions[letters[4]].width = ref_widths[3]
    ws.column_dimensions[letters[5]].width = ref_widths[4]

    # Remplir descriptions + styles colonne C
    for r in range(1, ws.max_row + 1):
        if not cell_can_write(ws, r, 3):
            continue
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        desc = description_for_row(c1, c2)
        if desc is not None:
            ws.cell(r, 3).value = desc

        if c1 == "Classe" and c2 == "Méthode":
            copy_cell_style(ws.cell(r, 4), ws.cell(r, 3))
        elif desc is not None and desc != "":
            copy_cell_style(ws.cell(r, 2), ws.cell(r, 3))
            al = ws.cell(r, 2).alignment
            ws.cell(r, 3).alignment = Alignment(
                horizontal=al.horizontal if al else None,
                vertical=al.vertical if al else "top",
                wrap_text=True,
            )

    insert_ui_section(ws)
    merge_responsable_column(ws)
    apply_reference_formatting(ws)
    apply_thin_borders(ws)

    try:
        wb.save(OUT)
        print(f"Written: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "._generated.xlsx")
        wb.save(alt)
        print(
            f"Could not write {OUT} (file open?). Wrote: {alt}",
        )
    print(f"Source (styles + structure): {ref_path}")


if __name__ == "__main__":
    main()
