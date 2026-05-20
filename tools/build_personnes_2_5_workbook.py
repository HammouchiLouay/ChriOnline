"""
Génère **cinq fichiers Excel séparés** (un par personne 1 à 5), même mise en forme
que la planification type : chaque classeur contient une ligne de **titre de poste**
en tête, puis colonnes Classe | Méthode | Description | Estimation | temps effectif | Responsable.

Réunion des cinq classeurs : couverture **complète** des classes Java du module serveur
(`chrionline-server/src/main/java`), avec chevauchement pédagogique possible (même classe
citée côté client et serveur si le périmètre diffère).

Sorties (répertoire parent du dossier tools/) :
  - Gestion des utilisateurs.xlsx
  - ChriOnline-Planning-Personne-2-Catalogue-et-stock.xlsx
  - ChriOnline-Planning-Personne-3-Panier-client.xlsx
  - ChriOnline-Planning-Personne-4-Commandes.xlsx
  - ChriOnline-Planning-Personne-5-Paiement-et-integration.xlsx

Exécution : python tools/build_personnes_2_5_workbook.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]

AUTHOR_PLACEHOLDER = "À compléter"

FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E1F2")
FILL_TITLE = PatternFill(fill_type="solid", fgColor="B4C6E7")
FILL_SECTION = PatternFill(fill_type="solid", fgColor="E2EFDA")
FILL_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_TITLE = Font(name="Calibri", size=12, bold=True)
THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MAX_COL = 6


def row_write(ws, r: int, values: tuple) -> None:
    for c, v in enumerate(values, start=1):
        ws.cell(r, c).value = v


def style_header(ws, r: int) -> None:
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(r, c)
        cell.fill = FILL_HEADER
        cell.font = FONT_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_job_title(ws, r: int) -> None:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
    cell = ws.cell(r, 1)
    cell.fill = FILL_TITLE
    cell.font = FONT_TITLE
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section_banner(ws, r: int) -> None:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
    cell = ws.cell(r, 1)
    cell.fill = FILL_SECTION
    cell.font = FONT_BOLD
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_data_row(ws, r: int, is_totale: bool = False) -> None:
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(r, c)
        cell.fill = FILL_SECTION if is_totale else FILL_WHITE
        cell.font = FONT_BOLD if is_totale else FONT_BODY
        if c in (1, 2, 3):
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        elif c in (4, 5):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_borders(ws) -> None:
    for r in range(1, ws.max_row + 1):
        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).border = THIN_BORDER


def set_widths(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


def _add_block(
    ws,
    job_title: str,
    source_label: str,
    fonctionnalites: list[tuple[str, str, float, float]],
    sections: list[tuple[str, list[tuple[str, str, str, float, float]]]],
) -> None:
    """Remplit la feuille : titre de poste (ligne 1), puis tableau."""
    r = 1
    ws.cell(r, 1).value = job_title
    style_job_title(ws, r)
    r += 2

    row_write(
        ws,
        r,
        ("Classe", "Méthode", "Description", "Estimation (h)", "temps effectif (h)", AUTHOR_PLACEHOLDER),
    )
    style_header(ws, r)
    r += 1

    style_section_banner(ws, r)
    ws.cell(r, 1).value = "Fonctionnalités (métier)"
    r += 1

    sum_f = 0.0
    sum_a = 0.0
    for libelle, desc, est, eff in fonctionnalites:
        row_write(ws, r, (None, libelle, desc, est, eff, AUTHOR_PLACEHOLDER))
        style_data_row(ws, r)
        sum_f += est
        sum_a += eff
        r += 1

    row_write(ws, r, ("Totale (fonctionnel)", None, None, round(sum_f, 2), round(sum_a, 2), AUTHOR_PLACEHOLDER))
    style_data_row(ws, r, is_totale=True)
    r += 1
    r += 1

    for section_title, methods in sections:
        style_section_banner(ws, r)
        ws.cell(r, 1).value = section_title
        r += 1
        row_write(
            ws,
            r,
            ("Classe", "Méthode", "Description", "Estimation (h)", "temps effectif (h)", AUTHOR_PLACEHOLDER),
        )
        style_header(ws, r)
        r += 1
        block_e = 0.0
        block_a = 0.0
        for classe, methode, desc, est, eff in methods:
            row_write(ws, r, (classe, methode, desc, est, eff, AUTHOR_PLACEHOLDER))
            style_data_row(ws, r)
            block_e += est
            block_a += eff
            r += 1
        row_write(ws, r, ("Totale", None, None, round(block_e, 2), round(block_a, 2), AUTHOR_PLACEHOLDER))
        style_data_row(ws, r, is_totale=True)
        r += 1
        r += 1

    style_section_banner(ws, r)
    ws.cell(r, 1).value = (
        f"Source : ChriOnline — {source_label} — aligné sur le code Java du dépôt "
        f"(docs/equipe/PERSONNE-0X-*.md)."
    )
    r += 1


# --- Données par personne (titre de poste + libellé source + contenu) ---

JOB_P1 = (
    "Titre du poste : responsable gestion des comptes, authentification et session — Personne 1 (ChriOnline)",
    "Personne 1 — Utilisateurs",
)


def data_personne1() -> tuple:
    fonc = [
        (
            "Inscription / connexion / déconnexion",
            "Comptes `user`, hachage BCrypt, jetons de session (`SessionRegistry`).",
            1.5,
            1.48,
        ),
        (
            "Profil, OTP, suppression de compte",
            "Mise à jour profil, codes e-mail, effacement transactionnel (`AccountDeletionDAO`).",
            1.2,
            1.18,
        ),
        (
            "E-mail transactionnel",
            "Configuration `MailConfigLoader` / `MailService` (SMTP, Resend, fallback console).",
            0.8,
            0.79,
        ),
    ]
    sections = [
        (
            "chrionline.User — modèle",
            [
                ("User", "constructeurs / getters / setters", "Bean utilisateur (id, hash, email, téléphone, rôle…).", 0.35, 0.36),
                ("User", "isEmailVerified()", "Accès au drapeau de vérification e-mail.", 0.08, 0.08),
            ],
        ),
        (
            "chrionline.UserDAO — persistance JDBC",
            [
                ("UserDAO", "createUser / findById / findByEmail / findByPhoneNumber", "CRUD ciblée sur la table `user`.", 0.55, 0.56),
                ("UserDAO", "emailExists / phoneNumberExists", "Unicité e-mail et téléphone.", 0.2, 0.21),
                ("UserDAO", "updatePassword / updateEmail / updatePhoneNumber / …", "Mises à jour de profil.", 0.35, 0.36),
                ("UserDAO", "deleteUser / updateEmailVerified", "Suppression et vérif e-mail.", 0.25, 0.26),
            ],
        ),
        (
            "chrionline.Authentification",
            [
                ("Authentification", "register(User)", "Inscription avec contrôle d’unicité.", 0.35, 0.36),
                ("Authentification", "loginByEmailOrPhone(...)", "Vérification BCrypt et session.", 0.45, 0.46),
            ],
        ),
        (
            "chrionline.BaseDonnees & chrionline.TestConnexion",
            [
                ("BaseDonnees", "getConnection()", "Pool / connexion MySQL (paramètres env).", 0.25, 0.26),
                ("BaseDonnees", "verifyConnection()", "Santé BDD (PING).", 0.1, 0.1),
                ("TestConnexion", "main / test JDBC", "Utilitaire de test de connexion locale.", 0.15, 0.14),
            ],
        ),
        (
            "chrionline.PhoneNumberLookup",
            [
                ("PhoneNumberLookup", "(normalisation / recherche)", "Aide à la résolution téléphone pour auth.", 0.2, 0.21),
            ],
        ),
        (
            "services.AuthService",
            [
                ("AuthService", "login(Message)", "LOGIN : JSON → Authentification → jeton session.", 0.55, 0.56),
                ("AuthService", "register(Message)", "REGISTER : création compte + e-mail vérif.", 0.45, 0.46),
            ],
        ),
        (
            "services.EmailVerificationService",
            [
                ("EmailVerificationService", "send(Message) / confirm(Message)", "EMAIL_VERIFY_* : codes et mise à jour base.", 0.5, 0.52),
                ("EmailVerificationService", "scheduleInitialVerificationEmail / clearPendingForUser", "Planification et nettoyage.", 0.25, 0.26),
            ],
        ),
        (
            "services.PasswordResetService",
            [
                ("PasswordResetService", "forgotPassword / resetPassword(Message)", "FORGOT_PASSWORD / RESET_PASSWORD.", 0.45, 0.46),
            ],
        ),
        (
            "services.ProfileSecurityService & services.ProfileService",
            [
                ("ProfileSecurityService", "sendOtp / verifyAndConsumeProfileOtp", "PROFILE_OTP_SEND et vérif OTP profil.", 0.4, 0.41),
                ("ProfileService", "updateProfile(Message)", "UPDATE_PROFILE : champs et mot de passe actuel.", 0.5, 0.52),
            ],
        ),
        (
            "persistence.AccountDeletionDAO & services.AccountDeletionService",
            [
                ("AccountDeletionDAO", "deleteAllForUser(Connection, int)", "DELETE ordonnés (FK) avant `user`.", 0.45, 0.46),
                ("AccountDeletionService", "deleteAccount(Message)", "DELETE_ACCOUNT : session + transaction.", 0.4, 0.41),
            ],
        ),
        (
            "services.MailConfigLoader & services.MailService",
            [
                ("MailConfigLoader", "load() / userConfigFilePath()", "Fusion properties + env.", 0.25, 0.26),
                ("MailService", "sendPlain / sendEmailVerification / sendPasswordResetCode / …", "Envoi e-mails et diagnostics.", 0.55, 0.56),
            ],
        ),
        (
            "server.SessionRegistry",
            [
                ("SessionRegistry", "issue / resolveUser / revoke", "Jetons opaque → userId (LOGOUT, GET_COMMANDES…).", 0.45, 0.46),
            ],
        ),
        (
            "common.PasswordHasher & common.MaskingUtil",
            [
                ("PasswordHasher", "hash / verify", "Encapsulation BCrypt côté app.", 0.2, 0.21),
                ("MaskingUtil", "(masquage chaînes sensibles)", "Affichage sécurisé (indices contact).", 0.15, 0.16),
            ],
        ),
        (
            "org.mindrot.jbcrypt.BCrypt",
            [
                ("BCrypt", "gensalt / hashpw / checkpw", "Bibliothèque tierce — hachage et vérif mots de passe.", 0.1, 0.1),
            ],
        ),
        (
            "common.ClientPrefs & common.ClientConfigLoader (client)",
            [
                ("ClientPrefs", "getString / putString / …", "Préférences locales (session persistée).", 0.25, 0.26),
                ("ClientConfigLoader", "load()", "Hôte / port client.", 0.15, 0.16),
            ],
        ),
        (
            "ui.ChriOnlineClientApp — authentification & compte (extrait)",
            [
                ("ChriOnlineClientApp", "submitSidebarLogin / openRegisterDialog / openForgotPasswordDialog", "Flux LOGIN / REGISTER / reset.", 0.55, 0.56),
                ("ChriOnlineClientApp", "applySession / persistSession / restorePersistedSession / logoutAccount", "Jeton session et préférences.", 0.5, 0.52),
                ("ChriOnlineClientApp", "sendAccountEmailVerification / submitProfileUpdateFromPage / confirmAndDeleteAccount", "Compte : vérif e-mail, profil, suppression.", 0.55, 0.54),
            ],
        ),
    ]
    return fonc, sections


JOB_P2 = (
    "Titre du poste : responsable catalogue, fiches produits et stock (Personne 2 — ChriOnline)",
    "Personne 2 — Produits",
)


def data_personne2() -> tuple:
    fonc = [
        (
            "Liste catalogue (pagination / catégorie)",
            "Consultation du catalogue MySQL avec filtre catégorie et pagination.",
            1.2,
            1.15,
        ),
        (
            "Détail produit",
            "Fiche produit (prix net, stock, image, marque) depuis la base.",
            0.6,
            0.58,
        ),
        (
            "Mise à jour du stock",
            "Message STOCK_UPDATE pour ajuster les quantités disponibles.",
            0.5,
            0.52,
        ),
        (
            "Soumissions vendeur / modération admin",
            "Fiches produit en attente, approbation, rejet (`ProductListingService`).",
            0.9,
            0.88,
        ),
        (
            "Catégories métier",
            "Liste des catégories pour le filtre UI (Tous en tête).",
            0.35,
            0.33,
        ),
    ]
    sections = [
        (
            "product.Product — DTO catalogue",
            [
                ("Product", "Product()", "Constructeur vide ; sérialisable pour socket.", 0.05, 0.05),
                ("Product", "getId / getName / getPrice / getStock / getImageUrl / …", "Accesseurs affichage, panier, commande.", 0.3, 0.31),
                ("Product", "setStock(int)", "Synchronise quantité côté objet après STOCK_UPDATE.", 0.1, 0.1),
            ],
        ),
        (
            "product.ProductListingInfo — record soumission",
            [
                ("ProductListingInfo", "champs record", "Métadonnées listing (statut, vendeur, SKU, rejet…).", 0.2, 0.21),
            ],
        ),
        (
            "product.ProductCatalogDAO — JDBC",
            [
                ("ProductCatalogDAO", "loadAll / loadByCategory", "SELECT catalogue complet ou filtré.", 0.4, 0.42),
                ("ProductCatalogDAO", "loadPage / loadPageByCategory", "Pagination SQL LIMIT/OFFSET.", 0.35, 0.36),
                ("ProductCatalogDAO", "findByProductId(int)", "Une ligne produit par id.", 0.2, 0.2),
                ("ProductCatalogDAO", "updateStock(int, int)", "UPDATE `products.stock`.", 0.25, 0.27),
            ],
        ),
        (
            "product.ProductRepository — façade",
            [
                ("ProductRepository", "getAll / getByCategory / getPage", "Façade JDBC + cache / fallback mémoire.", 0.45, 0.46),
                ("ProductRepository", "findById(String)", "Optional<Product> pour UI et `CommandeService`.", 0.28, 0.29),
                ("ProductRepository", "distinctCategories / updateStock", "Catégories et stock.", 0.3, 0.31),
            ],
        ),
        (
            "services.ProductService — messages socket",
            [
                ("ProductService", "list(Message)", "PRODUCT_LIST — liste sérialisée Base64.", 0.55, 0.56),
                ("ProductService", "categories(Message)", "PRODUCT_CATEGORIES — JSON tableau de chaînes.", 0.3, 0.31),
                ("ProductService", "details(Message)", "PRODUCT_DETAILS — un `Product`.", 0.35, 0.36),
                ("ProductService", "updateStock(Message)", "STOCK_UPDATE — id + stock.", 0.35, 0.38),
            ],
        ),
        (
            "services.ProductListingService — fiches vendeur / admin",
            [
                ("ProductListingService", "submit(Message)", "SUBMIT_PRODUCT_LISTING.", 0.45, 0.46),
                ("ProductListingService", "listPending / listMine", "LIST_PENDING_PRODUCTS / LIST_MY_PRODUCT_LISTINGS.", 0.4, 0.41),
                ("ProductListingService", "approve / reject", "APPROVE_PRODUCT_LISTING / REJECT_PRODUCT_LISTING.", 0.4, 0.41),
            ],
        ),
        (
            "common.TextUiNormalizer",
            [
                ("TextUiNormalizer", "normalizeFrenchUi / categoryMatchVariants", "Normalisation libellés UI ↔ SQL.", 0.28, 0.29),
            ],
        ),
        (
            "ui.ProductImageLoader & ui.BrandIconUtil (client)",
            [
                ("ProductImageLoader", "loadAsync(...)", "Chargement images distantes (.webp, cache).", 0.35, 0.34),
                ("BrandIconUtil", "(icônes marque)", "Ressources graphiques marques pour grilles UI.", 0.15, 0.16),
            ],
        ),
    ]
    return fonc, sections


JOB_P3 = (
    "Titre du poste : responsable du panier d’achat et de l’expérience d’achat (client JavaFX) — Personne 3",
    "Personne 3 — Panier",
)


def data_personne3() -> tuple:
    fonc = [
        ("Ajout au panier", "Map produit id → quantité (LinkedHashMap) ; pas de table SQL dédiée.", 0.8, 0.82),
        ("Total panier", "Somme prix × quantité depuis le catalogue chargé en mémoire.", 0.45, 0.44),
        (
            "Conversion en commande",
            "Chaîne `id:qty;…` → `CREATE_COMMANDE` ; prix figés côté serveur (`CommandeService`).",
            0.65,
            0.66,
        ),
        ("Vidage après commande", "cart.clear() après succès CREATE_COMMANDE.", 0.2, 0.22),
        ("Annulation côté UI", "Bouton annuler commande EN_ATTENTE → ANNULER_COMMANDE (session).", 0.35, 0.34),
    ]
    sections = [
        (
            "ui.ChriOnlineClientApp — panier, navigation, commandes UI",
            [
                ("ChriOnlineClientApp", "cart (LinkedHashMap<String,Integer>)", "État panier uniquement client.", 0.15, 0.15),
                ("ChriOnlineClientApp", "addProductDetailToCart (grille catalogue)", "Ajout / incrément quantités panier.", 0.4, 0.41),
                ("ChriOnlineClientApp", "updateCartSummary", "Récapitulatif montants et compteur articles.", 0.45, 0.46),
                ("ChriOnlineClientApp", "findProduct(String id)", "Résout prix et libellés depuis caches produits.", 0.3, 0.31),
                ("ChriOnlineClientApp", "buildProduitsPayload()", "Sérialise id:qty;… pour CREATE_COMMANDE.", 0.35, 0.36),
                ("ChriOnlineClientApp", "createCommandeFromCart()", "Envoie CREATE_COMMANDE ; vide le panier si OK.", 0.5, 0.52),
                ("ChriOnlineClientApp", "buildOrderCard / requestCancelCommande", "Cartes commande ; annulation EN_ATTENTE.", 0.45, 0.46),
                ("ChriOnlineClientApp", "refreshCommandes / rebuildOrdersCards", "GET_COMMANDES + affichage.", 0.4, 0.41),
            ],
        ),
        (
            "services.CommandeService — chaîne d’achat (côté serveur, lien panier)",
            [
                ("CommandeService", "createCommandeAvecProduits(userId, produitsData)", "Parse id:qty, `ProductRepository.findById`, insert.", 0.65, 0.66),
                ("CommandeService", "createCommande(userId, total)", "Variante ligne synthétique « Commande globale ».", 0.25, 0.26),
            ],
        ),
        (
            "server.RequestRouter (messages panier → commande)",
            [
                ("RequestRouter", "case CREATE_COMMANDE", "JSON userId + produits → CommandeService.", 0.35, 0.36),
                ("RequestRouter", "case ANNULER_COMMANDE", "JSON sessionToken + commandeId ; propriété commande.", 0.35, 0.36),
            ],
        ),
        (
            "ui.SocketApiClient — client TCP",
            [
                ("SocketApiClient", "send(Message)", "Requête une ligne JSON / réponse.", 0.25, 0.26),
                ("SocketApiClient", "CommandeFull / OrderLineSnapshot / parseCommandesFull", "Parse JSON commandes pour l’UI.", 0.45, 0.46),
                ("SocketApiClient", "parseCommandeSummaries / fetchProductList / fetchProductDetails", "Helpers catalogue et historique.", 0.35, 0.36),
            ],
        ),
        (
            "ui.UiMessages & ui.StageResizeBehavior & ui.LanDiscoveryClient",
            [
                ("UiMessages", "errorCode(String)", "Libellés français des codes serveur.", 0.2, 0.21),
                ("StageResizeBehavior", "(comportement redimensionnement)", "Ajustements JavaFX plein écran / ratio.", 0.2, 0.19),
                ("LanDiscoveryClient", "(découverte serveur LAN)", "UDP / découverte hôte pour le client.", 0.35, 0.36),
            ],
        ),
        (
            "server.NetworkInfo & server.LanDiscoveryProtocol & server.PublicIpHint",
            [
                ("NetworkInfo", "(adresses locales / bind)", "Aide configuration réseau serveur ↔ client.", 0.25, 0.26),
                ("LanDiscoveryProtocol", "(format paquets découverte)", "Protocole annonce serveur.", 0.2, 0.21),
                ("PublicIpHint", "(indication IP publique)", "Option WAN / affichage.", 0.15, 0.16),
            ],
        ),
    ]
    return fonc, sections


JOB_P4 = (
    "Titre du poste : responsable des commandes, statuts et persistance (serveur & BDD) — Personne 4",
    "Personne 4 — Commandes",
)


def data_personne4() -> tuple:
    fonc = [
        ("Création commande", "Depuis panier ou total simplifié ; lignes figées en `order_lines`.", 1.0, 1.05),
        ("Validation / annulation", "Statuts SQL : EN_ATTENTE → VALIDE ; ANNULEE si non VALIDE/PAYEE.", 0.65, 0.64),
        ("Historique utilisateur", "GET_COMMANDES avec jeton session ; JSON `models.Commande#toJson`.", 0.65, 0.67),
        ("Protocole socket texte", "Messages `common.Message` + JSON `JsonUtil` sur TCP.", 0.5, 0.48),
    ]
    sections = [
        (
            "common.Message & common.JsonUtil",
            [
                ("Message", "getType / getPayload / request(...)", "Ligne protocole client ↔ serveur.", 0.3, 0.31),
                ("JsonUtil", "toJson / toMap", "Sérialisation légère des payloads.", 0.25, 0.26),
            ],
        ),
        (
            "models.LigneCommande",
            [
                ("LigneCommande", "LigneCommande(...)", "Ligne figée (produitId, nom, qté, prix unitaire).", 0.2, 0.2),
                ("LigneCommande", "calculerSousTotal() / toJson()", "Sous-total et fragment JSON.", 0.2, 0.21),
            ],
        ),
        (
            "models.Commande",
            [
                ("Commande", "Commande(int, int)", "Statut initial EN_ATTENTE ; liste de lignes.", 0.25, 0.26),
                ("Commande", "ajouterLigne / calculerTotal / toJson()", "Montage et sérialisation pour GET_COMMANDES.", 0.4, 0.41),
            ],
        ),
        (
            "services.CommandeService",
            [
                ("CommandeService", "createCommande / createCommandeAvecProduits", "Insert via DAO + catalogue.", 0.65, 0.66),
                ("CommandeService", "validerCommande / annulerCommande", "Règles EN_ATTENTE / exclusions VALIDE-PAYEE.", 0.4, 0.41),
                ("CommandeService", "getCommandesByUser / getCommandeById", "Lecture historique et paiement.", 0.45, 0.46),
                ("CommandeService", "updateCommandeStatus", "Après simulation paiement côté ChriOnline.", 0.3, 0.31),
            ],
        ),
        (
            "persistence.CommandeDAO",
            [
                ("CommandeDAO", "insert(Connection, Commande)", "Transaction `orders` + `order_lines`.", 0.55, 0.56),
                ("CommandeDAO", "findById / findByUserId", "SELECT + lignes (JOIN produits pour noms).", 0.45, 0.46),
                ("CommandeDAO", "valider / annuler / updateStatus", "UPDATE `orders.status` (ANNULEE, PAYEE, …).", 0.4, 0.41),
            ],
        ),
        (
            "server.RequestRouter — routage principal",
            [
                ("RequestRouter", "route(Message)", "Switch sur type : auth, produits, commandes, paiement.", 0.55, 0.56),
                ("RequestRouter", "CREATE_COMMANDE / VALIDER_COMMANDE / ANNULER_COMMANDE", "Payload JSON ou id ; annulation avec session.", 0.45, 0.46),
                ("RequestRouter", "GET_COMMANDES", "sessionToken → userId → liste commandes.", 0.35, 0.38),
            ],
        ),
        (
            "server.ServerMain & server.ClientHandler",
            [
                ("ServerMain", "main", "Lance `ServerSocket`, boucle accept.", 0.3, 0.31),
                ("ClientHandler", "run()", "Une ligne JSON par socket ; RequestRouter.route.", 0.4, 0.41),
            ],
        ),
        (
            "client.TestClient",
            [
                ("TestClient", "main", "Menu console : tous les types de messages (dont commandes).", 0.35, 0.36),
            ],
        ),
    ]
    return fonc, sections


JOB_P5 = (
    "Titre du poste : responsable paiement simulé, historique et moyens enregistrés — Personne 5",
    "Personne 5 — Paiement",
)


def data_personne5() -> tuple:
    fonc = [
        ("Simulation paiement", "Module `ecommerce.personne5` — types, wallet, coupons, score fraude.", 1.3, 1.28),
        ("Pont socket ChriOnline", "`SocketPaymentService` met à jour `orders.status` et journal SQL.", 0.75, 0.76),
        ("Historique & moyens enregistrés", "`historique_paiement`, `methode_paiement_enregistree` (masqué).", 0.55, 0.54),
        ("Promotions modèle partagé", "`model.PromotionCampaign` (campagnes locales / doc).", 0.15, 0.14),
    ]
    sections = [
        (
            "ecommerce.personne5.service — cœur métier",
            [
                ("PaiementService", "simulerPaiement / confirmerPaiement / rembourserPaiement", "Flux carte, livraison, wallet, audit.", 0.95, 0.96),
                ("PaiementService", "verifierCoupon / calculerFraisLivraison", "Promo et livraison.", 0.45, 0.46),
                ("WalletService", "débit / crédit (wallet simulé)", "Solde utilisateur pour paiement.", 0.25, 0.26),
                ("CurrencyService", "conversion / arrondi", "Montants multi-devises.", 0.2, 0.21),
                ("AuditService", "enregistrerAction", "Journal métier simulation.", 0.2, 0.21),
                ("PromotionService", "règles promos", "Coupons et campagnes côté module.", 0.25, 0.26),
                ("FlashSaleService", "fenêtres promo", "Ventes flash simulées.", 0.2, 0.2),
                ("FraudDetectionService", "score fraude", "Heuristiques refus.", 0.25, 0.26),
                ("RiskAnalysisService", "analyse risque", "Complément scoring.", 0.2, 0.2),
                ("PaymentTimeoutService", "expiration tentative", "Timeouts paiement.", 0.15, 0.16),
                ("PaymentAnalyticsService", "stats paiement", "Agrégats internes.", 0.15, 0.16),
                ("PaymentRecommendationService", "recommandation mode paiement", "Suggestion UI/console.", 0.15, 0.15),
                ("CashbackService", "remises différées", "Cashback simulé.", 0.15, 0.16),
                ("LoyaltyService", "fidélité", "Points / paliers.", 0.15, 0.15),
                ("RetryPaymentService", "nouvelle tentative", "Rejeu après échec.", 0.15, 0.15),
                ("PaiementQueueService", "file d’attente", "Orchestration async simplifiée.", 0.15, 0.15),
            ],
        ),
        (
            "ecommerce.personne5.model — entités simulation",
            [
                ("Paiement", "champs / getters", "Montant, statut, type, frais, fraude…", 0.25, 0.26),
                ("Commande (personne5)", "modèle simulation", "Distinct de `models.Commande` (id String).", 0.2, 0.21),
                ("TypePaiement / StatutPaiement / StatutCommande", "énumérations", "Valeurs métier paiement.", 0.15, 0.15),
                ("Coupon / AuditLog / Wallet / PaiementStats", "DTO module", "Structures auxiliaires.", 0.25, 0.26),
            ],
        ),
        (
            "ecommerce.personne5.utils & config",
            [
                ("LoggerUtil", "logs structurés", "Traces module.", 0.1, 0.1),
                ("PaiementValidator / PaiementSecurityUtil", "contrôles entrées", "Validation et sécurité légère.", 0.2, 0.21),
                ("RecuGenerator", "reçu texte", "Reçu simulé.", 0.1, 0.1),
                ("NotificationUtil", "notifications", "Messages succès/échec.", 0.1, 0.1),
                ("TokenUtil / IdGeneratorUtil", "jetons / ids", "Identifiants corrélés.", 0.1, 0.1),
                ("AppConfig / BusinessRules", "constantes", "Paramètres module.", 0.15, 0.16),
            ],
        ),
        (
            "ecommerce.personne5.dashboard & main",
            [
                ("AdminDashboard", "console admin module", "Vue agrégée (hors client JavaFX principal).", 0.2, 0.19),
                ("PaiementApp", "main", "Point d’entrée autonome du sous-projet paiement.", 0.15, 0.15),
            ],
        ),
        (
            "services.SocketPaymentService",
            [
                ("SocketPaymentService", "simulatePayment(Message)", "SIMULATE_PAYMENT : JSON → PaiementService → DAO.", 0.8, 0.81),
            ],
        ),
        (
            "services.SavedPaymentService",
            [
                ("SavedPaymentService", "list(Message) / delete(Message)", "Moyens enregistrés : liste et suppression.", 0.45, 0.46),
            ],
        ),
        (
            "persistence.PaymentHistoryDAO & SavedPaymentMethodDAO",
            [
                ("PaymentHistoryDAO", "insert / …", "Historique des paiements ChriOnline.", 0.35, 0.36),
                ("SavedPaymentMethodDAO", "insert / listByUser / deleteForUser", "Cartes masquées persistées.", 0.4, 0.41),
            ],
        ),
        (
            "model.PromotionCampaign (racine model)",
            [
                ("PromotionCampaign", "constructeur / accesseurs", "Campagne promo (dates, remise) — usage documentaire / extension.", 0.15, 0.14),
            ],
        ),
    ]
    return fonc, sections


OUTPUTS: list[tuple[str, str, tuple, callable]] = [
    (
        "Gestion des utilisateurs.xlsx",
        "Planning P1",
        JOB_P1,
        data_personne1,
    ),
    (
        "ChriOnline-Planning-Personne-2-Catalogue-et-stock.xlsx",
        "Planning P2",
        JOB_P2,
        data_personne2,
    ),
    (
        "ChriOnline-Planning-Personne-3-Panier-client.xlsx",
        "Planning P3",
        JOB_P3,
        data_personne3,
    ),
    (
        "ChriOnline-Planning-Personne-4-Commandes.xlsx",
        "Planning P4",
        JOB_P4,
        data_personne4,
    ),
    (
        "ChriOnline-Planning-Personne-5-Paiement-et-integration.xlsx",
        "Planning P5",
        JOB_P5,
        data_personne5,
    ),
]


def main() -> None:
    for filename, sheet_name, job_info, data_fn in OUTPUTS:
        job_title, source_label = job_info
        fonc, sections = data_fn()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        _add_block(ws, job_title, source_label, fonc, sections)
        set_widths(ws)
        apply_borders(ws)
        path = ROOT / filename
        wb.save(path)
        print(f"Written: {path}")


if __name__ == "__main__":
    main()
